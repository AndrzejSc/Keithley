from openpyxl import load_workbook
import pandas as pd


class AppendToExcel():
    def __init__(self, filename, sheet_name='Sheet1', startrow=None, truncate_sheet=False, columns=None, **to_excel_kwargs):
        self.filename = filename
        self.sheet_name = sheet_name
        self.startrow = startrow
        self.truncate_sheet = truncate_sheet
        self.columns = columns
        # print(self.columns)
        self.to_excel_kwargs = to_excel_kwargs
        self.dataToAdd = None
        self.writer = pd.ExcelWriter(filename, engine='openpyxl')

    def append(self, df, startrow=None, header=True):
        self.startrow = startrow
        self.dataToAdd = pd.DataFrame(df)
        self.dataToAdd = df

        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
    
        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]
    
        Returns: None
        """

        # ignore [engine] parameter if it was passed
        if 'engine' in self.to_excel_kwargs:
            self.to_excel_kwargs.pop('engine')

        try:
            # try to open an existing workbook
            self.writer.book = load_workbook(self.filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and self.sheet_name in self.writer.book.sheetnames:
                startrow = self.writer.book[self.sheet_name].max_row

            # truncate sheet
            if self.truncate_sheet and self.sheet_name in self.writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = self.writer.book.sheetnames.index(self.sheet_name)
                # remove [sheet_name]
                self.writer.book.remove(self.writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                self.writer.book.create_sheet(self.sheet_name, idx)

            # copy existing sheets
            self.writer.sheets = {ws.title: ws for ws in self.writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            print("Nie udało się znaleźć XXX. Z klasy AppendToExcel")
            pass

        if startrow is None:
            startrow = 0

        if header:
            self.dataToAdd.to_excel(self.writer, self.sheet_name, startrow=startrow, columns=self.columns)
        else:
            self.dataToAdd.to_excel(self.writer, self.sheet_name, startrow=startrow, header=False)
        # write out the new sheet
        self.writer.save()
